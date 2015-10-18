'''
Leaseweb CSV Invoice Converter
V1.0
Dimofinf, Inc
'''

import csv
import xlsxwriter
import os
from openpyxl import load_workbook
from functions import list_duplicates
from functions import replace_digits

# CSV filename
lw_csv = input("Please type the CSV filename: ")

# TEMP File name
export_temp = "invoice_tmp.xlsx"

# Final file delivered
export_invoice = "invoice.xlsx"

# Open XLSX file and adding sheets
workbook = xlsxwriter.Workbook(export_invoice)
# Formatting CELLS
cells_titles_format = workbook.add_format({'bold': True})
cells_titles_format.set_bg_color("#81BEF7")

cells_titles_format_report = workbook.add_format({'bold': True})
cells_titles_format_report.set_bg_color("#00FF00")

cells_titles_format_report_results = workbook.add_format({'bold': True})
cells_titles_format_report_results.set_bg_color("#F7FE2E")

# Adding SpreadSheets
worksheet_Servers_Pricing = workbook.add_worksheet("Servers_Pricing")
worksheet_VPS = workbook.add_worksheet("VPS")
worksheet_Windows_Licenses = workbook.add_worksheet("Windows_Licenses")
worksheet_Plesk_Licenses = workbook.add_worksheet("Plesk_Licenses")
worksheet_cPanel_licenses = workbook.add_worksheet("cPanel_licenses")
worksheet_Cabling = workbook.add_worksheet("Cabling")
worksheet_Extras = workbook.add_worksheet("Extras")
worksheet_Reports = workbook.add_worksheet("Reports")

# Creating tmp file to save serverIDs
workbook_tmp = xlsxwriter.Workbook(export_temp)
worksheet_tmp_Servers_Pricing = workbook_tmp.add_worksheet("Servers_Pricing")
worksheet_tmp_VPS = workbook_tmp.add_worksheet("VPS")

# Declaring Awesome variables for columns
serverid_col = 0
uplink_col = 1
kvm_col = 2
# Server Column is the price of the server in a separate column
server_col = 3
rackspace_col = 4
traffic_col = 5
ip_col = 6
apc_col = 7
ipmi_col = 8
support_col = 8
switch_col = 9

# Declaring Total variables
total_price_server = 10
total_price_sum_server = 11
total_price_sum_vps = 6
total_price_sum_windows = 3
total_price_sum_plesk = 3
total_price_sum_cpanel = 3
total_price_sum_cabling = 3
total_price_sum_extras = 3

# Declaring VPS variables for columns
serverid_col_vps = 0
server_col_vps = 1
traffic_col_vps = 2
support_col_vps = 3
total_price_vps = 4

# Format to write row, column, content
## Servers_Pricing Sheet
### Set width for column
worksheet_Servers_Pricing.set_column(0, serverid_col, 21)
worksheet_Servers_Pricing.set_column(rackspace_col, rackspace_col, 10)
worksheet_Servers_Pricing.set_column(total_price_server, total_price_server, 16)
worksheet_Servers_Pricing.set_column(total_price_sum_server, total_price_sum_server, 16)

worksheet_Servers_Pricing.write(0, uplink_col, "Uplink", cells_titles_format)
worksheet_Servers_Pricing.write(0, kvm_col, "KVM", cells_titles_format)
worksheet_Servers_Pricing.write(0, server_col, "Server", cells_titles_format)
worksheet_Servers_Pricing.write(0, rackspace_col, "RackSpace", cells_titles_format)
worksheet_Servers_Pricing.write(0, traffic_col, "Traffic", cells_titles_format)
worksheet_Servers_Pricing.write(0, ip_col, "IP", cells_titles_format)
worksheet_Servers_Pricing.write(0, apc_col, "APC", cells_titles_format)
worksheet_Servers_Pricing.write(0, ipmi_col, "IPMI", cells_titles_format)
worksheet_Servers_Pricing.write(0, support_col, "Support", cells_titles_format)
worksheet_Servers_Pricing.write(0, switch_col, "Switches", cells_titles_format)
worksheet_Servers_Pricing.write(0, total_price_server, "Server Price Euro", cells_titles_format)
worksheet_Servers_Pricing.write(6, total_price_sum_server, "Total Sheet Price", cells_titles_format)


## VPS Sheet
### Set width for column
worksheet_VPS.set_column(0, serverid_col, 21)
worksheet_VPS.set_column(total_price_vps, total_price_vps, 16)
worksheet_VPS.set_column(total_price_sum_vps, total_price_sum_vps, 16)
worksheet_VPS.write(0, server_col_vps, "Server", cells_titles_format)
worksheet_VPS.write(0, traffic_col_vps, "Traffic", cells_titles_format)
worksheet_VPS.write(0, support_col_vps, "Support", cells_titles_format)
worksheet_VPS.write(0, total_price_vps, "VPS Price Euro", cells_titles_format)
worksheet_VPS.write(6, total_price_sum_vps, "Total Sheet Price", cells_titles_format)

## cPanel_licenses Sheet
### Set width for column
worksheet_cPanel_licenses.set_column(0, 0, 40)
worksheet_cPanel_licenses.set_column(total_price_sum_cpanel, total_price_sum_cpanel, 16)
worksheet_cPanel_licenses.write(0, 0, "License Name", cells_titles_format)
worksheet_cPanel_licenses.write(0, 1, "Amount", cells_titles_format)
worksheet_cPanel_licenses.write(6, total_price_sum_cpanel, "Total Sheet Price", cells_titles_format)

## Plesk_Licenses Sheet
### Set width for column
worksheet_Plesk_Licenses.set_column(0, 0, 40)
worksheet_Plesk_Licenses.set_column(total_price_sum_plesk, total_price_sum_plesk, 16)
worksheet_Plesk_Licenses.write(0, 0, "License Name", cells_titles_format)
worksheet_Plesk_Licenses.write(0, 1, "Amount", cells_titles_format)
worksheet_Plesk_Licenses.write(6, total_price_sum_plesk, "Total Sheet Price", cells_titles_format)

## Extras Sheet
### Set width for column
worksheet_Extras.set_column(0, 0, 40)
worksheet_Extras.set_column(total_price_sum_extras, total_price_sum_extras, 16)
worksheet_Extras.write(0, 0, "Name", cells_titles_format)
worksheet_Extras.write(0, 1, "Amount", cells_titles_format)
worksheet_Extras.write(6, total_price_sum_extras, "Total Sheet Price", cells_titles_format)

## Windows_Licenses Sheet
### Set width for column
worksheet_Windows_Licenses.set_column(0, 0, 83)
worksheet_Windows_Licenses.set_column(total_price_sum_windows, total_price_sum_windows, 16)
worksheet_Windows_Licenses.write(0, 0, "Name", cells_titles_format)
worksheet_Windows_Licenses.write(0, 1, "Amount", cells_titles_format)
worksheet_Windows_Licenses.write(6, total_price_sum_windows, "Total Sheet Price", cells_titles_format)

## Cabling Sheet
### Set width for column
worksheet_Cabling.set_column(0, 0, 40)
worksheet_Cabling.set_column(total_price_sum_cabling, total_price_sum_cabling, 16)
worksheet_Cabling.write(0, 0, "Name", cells_titles_format)
worksheet_Cabling.write(0, 1, "Amount", cells_titles_format)
worksheet_Cabling.write(6, total_price_sum_cabling, "Total Sheet Price", cells_titles_format)

## Reports Sheet
### Set width for column
number_of_cpanel = 0
number_of_plesk = 1
number_of_windows = 2
number_of_servers = 3
number_of_vps = 4
total_cpanel_report = 7
total_plesk_report = 8
total_windows_report = 9
total_servers_report = 10
total_vps_report = 11
total_cabling_report = 12
total_extras_report = 13
total_report_sum = 16
total_invoice_csv = 17

# Modify the width of Column A
worksheet_Reports.set_column(0, 0, 26)

# Modify the width of Column D
worksheet_Reports.set_column(3, 3, 26)

worksheet_Reports.write(number_of_cpanel, 0, "Number of cPanel licenses", cells_titles_format)
worksheet_Reports.write(number_of_plesk, 0, "Number of Plesk licenses", cells_titles_format)
worksheet_Reports.write(number_of_windows, 0, "Number of Windows licenses", cells_titles_format)
worksheet_Reports.write(number_of_servers, 0, "Number of Servers", cells_titles_format)
worksheet_Reports.write(number_of_vps, 0, "Number of VPS", cells_titles_format)
worksheet_Reports.write(total_cpanel_report, 0, "Total cPanel Licenses Cost", cells_titles_format)
worksheet_Reports.write(total_plesk_report, 0, "Total Plesk Licenses Cost", cells_titles_format)
worksheet_Reports.write(total_windows_report, 0, "Total Windows Licenses Cost", cells_titles_format)
worksheet_Reports.write(total_servers_report, 0, "Total Servers Cost", cells_titles_format)
worksheet_Reports.write(total_vps_report, 0, "Total VPS Cost", cells_titles_format)
worksheet_Reports.write(total_cabling_report, 0, "Total Cabling", cells_titles_format)
worksheet_Reports.write(total_extras_report, 0, "Total Extra", cells_titles_format)
worksheet_Reports.write(total_report_sum, 0, "Total Invoice from XLSX", cells_titles_format_report)
worksheet_Reports.write(total_invoice_csv, 0, "Total Invoice from CSV", cells_titles_format_report)
worksheet_Reports.write(0, 3, "Importing Results", cells_titles_format_report_results)

# Open csv file in read mode
with open(lw_csv, "r") as lwinvoice:
    # Read the file and store contents in variable
    invoice_data = csv.reader(lwinvoice, delimiter=";")

    # Declaring variables used in loops as counters
    row_counter_cpanel = 1
    row_counter_extras = 1
    row_counter_windows_licenses = 1
    row_counter_plesk = 1
    row_counter_cabling = 1
    row_counter_server = 1
    row_counter_vps = 1
    global_counter = 0
    global_counter_report = 0
    # Start getting contents from the csv file row-by-row
    for row in invoice_data:
        # We should expect index errors, Let's keep it away
        try:
            # Search for CSV Amount and write to the corresponded sheet
            if "Amount" in row[0]:
                worksheet_Reports.write(total_invoice_csv, 1, float(row[1]))

            # contents of row are stored in list, Let's filter what we need
            ## Store the item type whether if it's Licenses, Serverhosting, Extras ..etc
            itemtype = row[0]
            ## Store the item name such the server id
            itemname = row[1]
            # The description for every product in text
            iteminfo = row[2]
            # Total Price for the item
            itemprice = row[6]

            global_counter += 1

            # Search for cPanel/WHM product and write to the corresponded sheet
            if "cPanel" in iteminfo and "Licenses" in itemtype:
                worksheet_cPanel_licenses.write(row_counter_cpanel, 0, iteminfo)
                worksheet_cPanel_licenses.write(row_counter_cpanel, 1, float(itemprice))
                row_counter_cpanel += 1

                # This counter for total items collected during the app execution
                global_counter_report += 1

            # Search for Plesk product and write to the corresponded sheet
            if "Plesk" in iteminfo and "Licenses" in itemtype:
                worksheet_Plesk_Licenses.write(row_counter_plesk, 0, iteminfo)
                worksheet_Plesk_Licenses.write(row_counter_plesk, 1, float(itemprice))
                row_counter_plesk += 1

                global_counter_report += 1

            # Search for windows licenses and write to the corresponded sheet
            if "Windows Server" in iteminfo or "SQL" in iteminfo and "Licenses" in itemtype:
                worksheet_Windows_Licenses.write(row_counter_windows_licenses, 0, iteminfo)
                worksheet_Windows_Licenses.write(row_counter_windows_licenses, 1, float(itemprice))
                row_counter_windows_licenses += 1

                global_counter_report += 1

            # Search for any extras and write to the corresponded sheet
            if "Extras" in itemtype:
                worksheet_Extras.write(row_counter_extras, 0, iteminfo)
                worksheet_Extras.write(row_counter_extras, 1, float(itemprice))
                row_counter_extras += 1

                global_counter_report += 1

            # Search for any cabling and write to the corresponded sheet
            if "Cabling" in itemtype:
                worksheet_Cabling.write(row_counter_cabling, 0, iteminfo)
                worksheet_Cabling.write(row_counter_cabling, 1, float(itemprice))
                row_counter_cabling += 1

                global_counter_report += 1

            # Search for serverID and write it in two separate columns, The first for reference\
            #  and the other for pricing
            if "Server:" in iteminfo and "Serverhosting" in itemtype:
                worksheet_Servers_Pricing.write(row_counter_server, server_col, float(itemprice))
                worksheet_Servers_Pricing.write(row_counter_server, serverid_col, itemname)
                # Save the same data in temp file
                worksheet_tmp_Servers_Pricing.write(row_counter_server, server_col, float(itemprice))
                worksheet_tmp_Servers_Pricing.write(row_counter_server, serverid_col, itemname)
                row_counter_server += 1

                global_counter_report += 1

            # Search for serverID - VPS type and write it in two separate columns, The first for reference \
            # and the other for pricing
            if "VPS" in itemtype and itemname in iteminfo:
                worksheet_VPS.write(row_counter_vps, server_col_vps, float(itemprice))
                worksheet_VPS.write(row_counter_vps, serverid_col_vps, itemname)
                # Save the same data in temp file
                worksheet_tmp_VPS.write(row_counter_vps, server_col_vps, float(itemprice))
                worksheet_tmp_VPS.write(row_counter_vps, serverid_col_vps, itemname)
                row_counter_vps += 1

                global_counter_report += 1
        except:
            pass
# Close the temp file and let openpyxl get the required data
workbook_tmp.close()

# Open the new created temp file
wb = load_workbook(export_temp)
# We need to get and store information in Servers_Pricing and VPS sheet, Let's get our data from it
server_pricing_sheet = wb['Servers_Pricing']
vps_pricing_sheet = wb['VPS']

# ReOpen the CSV file to store the other data missing in Servers_Pricing sheet
with open(lw_csv, "r") as lwinvoice:
    invoice_data = csv.reader(lwinvoice, delimiter=";")

    # Well, We have the serverID reference columns ready, We need to get the position of each server to use later
    def get_server_position(serverid):
        servers = []
        # Loop through the sheet in the server counter range range, ADD 1 to match the last server value
        for i in range(2, row_counter_server+1):
            # Get the column ID from this loop
            server_name = server_pricing_sheet['A'+str(i)].value
            # Append the value to an awesome LIST
            servers.append(server_name)
        # GET the index by the serverID and return it to the user
        server_index = servers.index(serverid)
        return server_index+1

    # Well, We have the serverID reference columns ready, We need to get the position of each server to use later
    def get_vps_position(serverid):
        vps = []
        # Loop through the sheet in the server counter range range, ADD 1 to match the last server value
        for i in range(2, row_counter_vps+1):
            # Get the column ID from this loop
            vps_name = vps_pricing_sheet['A'+str(i)].value
            # Append the value to an awesome LIST
            vps.append(vps_name)
        # GET the index by the serverID and return it to the user
        vps_index = vps.index(serverid)
        return vps_index+1

    # Loop again through CSV data row-by-row to get server info
    # Define empty list to store ALL APC server name
    apc_list_names = []
    # Define empty list to store ALL APC values
    apc_list_values = []
    # Define dictionry to store APC servername and value based on a postfix ( This is used for duplication purposes )
    apc_dict = {}
    apc_counter = 0

    # Define empty list to store ALL uplink server name
    uplink_list_names = []
    # Define empty list to store ALL uplink values
    uplink_list_values = []
    # Define dictionry to store uplink servername and value based on a postfix ( This is used for duplication purposes )
    uplink_dict = {}
    uplink_counter = 0

    # Define empty list to store ALL ip server name
    ip_list_names = []
    # Define empty list to store ALL ip values
    ip_list_values = []
    # Define dictionry to store ip servername and value based on a postfix ( This is used for duplication purposes )
    ip_dict = {}
    ip_counter = 0

    # Define empty list to store ALL switch server name
    switch_list_names = []
    # Define empty list to store ALL switch values
    switch_list_values = []
    # Define dictionry to store switch servername and value based on a postfix ( This is used for duplication purposes )
    switch_dict = {}
    switch_counter = 0
    for row in invoice_data:
        try:
            itemtype = row[0]
            itemname = row[1]
            # The description for every product in text
            iteminfo = row[2]
            itemprice = row[6]

            if "APC" in iteminfo and "Serverhosting" in itemtype:
                apc_list_names.append(itemname)
                apc_list_values.append(itemprice)
                # Store itemname and itemprice with postfix for duplication issue
                apc_dict[itemname+"dimofinf%d" % apc_counter] = itemprice
                apc_counter += 1
                worksheet_Servers_Pricing.write(get_server_position(itemname), apc_col, float(itemprice))

                global_counter_report += 1

            uplink_names = ['GE PORT', 'CONNECTIVITY']
            for uplink in uplink_names:
                if "Serverhosting" in itemtype and uplink in iteminfo.upper():
                    uplink_list_names.append(itemname)
                    uplink_list_values.append(itemprice)
                    # Store itemname and itemprice with postfix for duplication issue
                    uplink_dict[itemname+"dimofinf%d" % uplink_counter] = itemprice
                    uplink_counter += 1
                    worksheet_Servers_Pricing.write(get_server_position(itemname), uplink_col, float(itemprice))

                    global_counter_report += 1

            if "KVM" in iteminfo and "Serverhosting" in itemtype:
                worksheet_Servers_Pricing.write(get_server_position(itemname), kvm_col, float(itemprice))

                global_counter_report += 1

            if "Rackspace" in iteminfo and "Serverhosting" in itemtype:
                worksheet_Servers_Pricing.write(get_server_position(itemname), rackspace_col, float(itemprice))

                global_counter_report += 1

            if "Datatraffic" in iteminfo or "Bandwidth" in iteminfo and "Serverhosting" in itemtype:
                worksheet_Servers_Pricing.write(get_server_position(itemname), traffic_col, float(itemprice))

                global_counter_report += 1

            ips_name = ['Extra IP', 'IP Announcement']
            for ip in ips_name:
                if "Serverhosting" in itemtype and ip in iteminfo:
                    worksheet_Servers_Pricing.write(get_server_position(itemname), ip_col, float(itemprice))

                    ip_list_names.append(itemname)
                    ip_list_values.append(itemprice)
                    # Store itemname and itemprice with postfix for duplication issue
                    ip_dict[itemname+"dimofinf%d" % ip_counter] = itemprice
                    ip_counter += 1

                    global_counter_report += 1

            support_pkgs = ['Basic', 'Bronze', 'Silver', 'Gold', 'Platinum']
            for sup_pkg in support_pkgs:
                if "Serverhosting" in itemtype and sup_pkg in iteminfo:
                    worksheet_Servers_Pricing.write(get_server_position(itemname), support_col, float(itemprice))

                    global_counter_report += 1

            switch_names = ['GE SWITCH']
            for switch in switch_names:
                if "Serverhosting" in itemtype and switch in iteminfo.upper():
                    switch_list_names.append(itemname)
                    switch_list_values.append(itemprice)
                    # Store itemname and itemprice with postfix for duplication issue
                    switch_dict[itemname+"dimofinf%d" % switch_counter] = itemprice
                    switch_counter += 1
                    worksheet_Servers_Pricing.write(get_server_position(itemname), switch_col, float(itemprice))

                    global_counter_report += 1
        except:
            pass

# Get a list of duplicated APC servernames ( clear and not duplicated values )
duplicate_apc = list_duplicates(apc_list_names)

# Define a duplicated list of values for both of names and price, I'm using it for indexing purposes and to merge later
duplicate_apc_names_list = []
duplicate_apc_values_list = []

# Loop through the dictionary
for apc_info in apc_dict.items():
    # Limit the world by looping only what you need "Duplicated APC info"
    for dups in duplicate_apc:
        apc_name = apc_info[0]
        apc_price = apc_info[1]
        if dups in apc_name:
            # Store the name and value the duplicated list
            duplicate_apc_names_list.append(replace_digits(apc_name))
            duplicate_apc_values_list.append(replace_digits(apc_price))

# Merge apc name and apc price together
merged_apc_list = zip(duplicate_apc_names_list, duplicate_apc_values_list)

# Define the final dictionary of APC which will store one key with multiple values, This is used to set one APC name
# with multiple values
final_apc_dict = dict()

# Loop through the merged list and set one key and multiple values, Dict output should be something like
# {'key1': [val1, val2], 'key2': [val1, val2]}

for line in merged_apc_list:
    if line[0] in final_apc_dict:
        # IF the key found, set the price
        final_apc_dict[line[0]].append(float(line[1]))
    else:
        # IF the key not found, create it
        final_apc_dict[line[0]] = [float(line[1])]

# Finally loop through our final dictionary info, Store the APC name and sum the values of duplicated!
for x, y in final_apc_dict.items():
    total = sum(y)
    worksheet_Servers_Pricing.write(get_server_position(x), apc_col, float(total))

# Get a list of duplicated ip servernames ( clear and not duplicated values )
duplicate_ip = list_duplicates(ip_list_names)

# Define a duplicated list of values for both of names and price, I'm using it for indexing purposes and to merge later
duplicate_ip_names_list = []
duplicate_ip_values_list = []

# Loop through the dictionary
for ip_info in ip_dict.items():
    # Limit the world by looping only what you need "Duplicated ip info"
    for dups in duplicate_ip:
        ip_name = ip_info[0]
        ip_price = ip_info[1]
        if dups in ip_name:
            # Store the name and value the duplicated list
            duplicate_ip_names_list.append(replace_digits(ip_name))
            duplicate_ip_values_list.append(replace_digits(ip_price))

# Merge ip name and ip price together
merged_ip_list = zip(duplicate_ip_names_list, duplicate_ip_values_list)

# Define the final dictionary of ip which will store one key with multiple values, This is used to set one ip name
# with multiple values
final_ip_dict = dict()

# Loop through the merged list and set one key and multiple values, Dict output should be something like
# {'key1': [val1, val2], 'key2': [val1, val2]}

for line in merged_ip_list:
    if line[0] in final_ip_dict:
        # IF the key found, set the price
        final_ip_dict[line[0]].append(float(line[1]))
    else:
        # IF the key not found, create it
        final_ip_dict[line[0]] = [float(line[1])]

# Finally loop through our final dictionary info, Store the ip name and sum the values of duplicated!
for x, y in final_ip_dict.items():
    total = sum(y)
    worksheet_Servers_Pricing.write(get_server_position(x), ip_col, float(total))

# Get a list of duplicated switch servernames ( clear and not duplicated values )
duplicate_switch = list_duplicates(switch_list_names)

# Define a duplicated list of values for both of names and price, I'm using it for indexing purposes and to merge later
duplicate_switch_names_list = []
duplicate_switch_values_list = []

# Loop through the dictionary
for switch_info in switch_dict.items():
    # Limit the world by looping only what you need "Duplicated switch info"
    for dups in duplicate_switch:
        switch_name = switch_info[0]
        switch_price = switch_info[1]
        if dups in switch_name:
            # Store the name and value the duplicated list
            duplicate_switch_names_list.append(replace_digits(switch_name))
            duplicate_switch_values_list.append(replace_digits(switch_price))

# Merge switch name and switch price together
merged_switch_list = zip(duplicate_switch_names_list, duplicate_switch_values_list)

# Define the final dictionary of switch which will store one key with multiple values, This is used to set one switch name
# with multiple values
final_switch_dict = dict()

# Loop through the merged list and set one key and multiple values, Dict output should be something like
# {'key1': [val1, val2], 'key2': [val1, val2]}

for line in merged_switch_list:
    if line[0] in final_switch_dict:
        # IF the key found, set the price
        final_switch_dict[line[0]].append(float(line[1]))
    else:
        # IF the key not found, create it
        final_switch_dict[line[0]] = [float(line[1])]

# Finally loop through our final dictionary info, Store the switch name and sum the values of duplicated!
for x, y in final_switch_dict.items():
    total = sum(y)
    worksheet_Servers_Pricing.write(get_server_position(x), switch_col, float(total))

# Get a list of duplicated uplink servernames ( clear and not duplicated values )
duplicate_uplink = list_duplicates(uplink_list_names)

# Define a duplicated list of values for both of names and price, I'm using it for indexing purposes and to merge later
duplicate_uplink_names_list = []
duplicate_uplink_values_list = []

# Loop through the dictionary
for uplink_info in uplink_dict.items():
    # Limit the world by looping only what you need "Duplicated uplink info"
    for dups in duplicate_uplink:
        uplink_name = uplink_info[0]
        uplink_price = uplink_info[1]
        if dups in uplink_name:
            # Store the name and value the duplicated list
            duplicate_uplink_names_list.append(replace_digits(uplink_name))
            duplicate_uplink_values_list.append(replace_digits(uplink_price))

# Merge uplink name and uplink price together
merged_uplink_list = zip(duplicate_uplink_names_list, duplicate_uplink_values_list)

# Define the final dictionary of uplink which will store one key with multiple values, This is used to set one uplink name
# with multiple values
final_uplink_dict = dict()

# Loop through the merged list and set one key and multiple values, Dict output should be something like
# {'key1': [val1, val2], 'key2': [val1, val2]}

for line in merged_uplink_list:
    if line[0] in final_uplink_dict:
        # IF the key found, set the price
        final_uplink_dict[line[0]].append(float(line[1]))
    else:
        # IF the key not found, create it
        final_uplink_dict[line[0]] = [float(line[1])]

# Finally loop through our final dictionary info, Store the uplink name and sum the values of duplicated!
for x, y in final_uplink_dict.items():
    total = sum(y)
    worksheet_Servers_Pricing.write(get_server_position(x), uplink_col, float(total))


# ReOpen the CSV file to store VPS Data in VPS sheet
with open(lw_csv, "r") as lwinvoice:
    invoice_data = csv.reader(lwinvoice, delimiter=";")
    # Loop again through CSV data row-by-row to get VPS info
    for row in invoice_data:
        try:
            itemtype = row[0]
            itemname = row[1]
            # The description for every product in text
            iteminfo = row[2]
            itemprice = row[6]

            if "Datatraffic" in iteminfo and "VPS" in itemtype:
                worksheet_VPS.write(get_vps_position(itemname), traffic_col_vps, float(itemprice))

                global_counter_report += 1

            support_pkgs = ['Basic', 'Bronze', 'Silver', 'Gold', 'Platinum']
            for sup_pkg in support_pkgs:
                if "VPS" in itemtype and sup_pkg in iteminfo:
                    worksheet_VPS.write(get_vps_position(itemname), support_col_vps, float(itemprice))

                    global_counter_report += 1

        except:
            pass

# Writing Sum formulas ( row, column , formula )
# Loop on the servers counter and type the sum function with a note that rows in xlsx library starts from zero
# And in Excel starts from 1 and the reference columns stars from 2, we should then add 1 to the servers_row counter
for servers_row in range(1, row_counter_server):
    worksheet_Servers_Pricing.write_formula(servers_row, total_price_server,
            '=SUM(B%d,C%d,D%d,E%d,F%d,G%d,H%d,I%d,J%d)' % (servers_row+1, servers_row+1, servers_row+1,
            servers_row+1, servers_row+1, servers_row+1, servers_row+1, servers_row+1, servers_row+1))

# Calculate the Total sum in Servers Pricing sheet ( row, col, content )
worksheet_Servers_Pricing.write_formula(7, total_price_sum_server,
            # GET the range to sum from the servers counter
            '=SUM(K2:K%d)' % row_counter_server)

# Loop through VPS and execute the sum function
for vps_row in range(1, row_counter_vps):
    worksheet_VPS.write_formula(vps_row, total_price_vps,
            '=SUM(B%d,C%d,D%d)' % (vps_row+1, vps_row+1, vps_row+1))

# Calculate the Total sum in VPS sheet
worksheet_VPS.write_formula(7, total_price_sum_vps,
            # GET the range to sum from the vps counter
            '=SUM(E2:E%d)' % row_counter_vps)

# Calculate the Total sum in Windows_Licenses sheet
worksheet_Windows_Licenses.write_formula(7, total_price_sum_windows,
            # GET the range to sum from the counter
            '=SUM(B2:B%d)' % row_counter_windows_licenses)

# Calculate the Total sum in Plesk_Licenses sheet
worksheet_Plesk_Licenses.write_formula(7, total_price_sum_plesk,
            # GET the range to sum from the counter
            '=SUM(B2:B%d)' % row_counter_plesk)

# Calculate the Total sum in cPanel_Licenses sheet
worksheet_cPanel_licenses.write_formula(7, total_price_sum_cpanel,
            # GET the range to sum from the counter
            '=SUM(B2:B%d)' % row_counter_cpanel)

# Calculate the Total sum in Cabling sheet
worksheet_Cabling.write_formula(7, total_price_sum_cabling,
            # GET the range to sum from the counter
            '=SUM(B2:B%d)' % row_counter_cabling)

# Calculate the Total sum in Extras sheet
worksheet_Extras.write_formula(7, total_price_sum_extras,
            # GET the range to sum from the counter
            '=SUM(B2:B%d)' % row_counter_extras)

# Calculate Number of cPanel licenses, Discount 1 because the items starts from A2
worksheet_Reports.write(number_of_cpanel, 1, row_counter_cpanel-1)

# Calculate Number of Plesk licenses, Discount 1 because the items starts from A2
worksheet_Reports.write(number_of_plesk, 1, row_counter_plesk-1)

# Calculate Number of Windows licenses, Discount 1 because the items starts from A2
worksheet_Reports.write(number_of_windows, 1, row_counter_windows_licenses-1)

# Calculate Number of Servers, Discount 1 because the items starts from A2
worksheet_Reports.write(number_of_servers, 1, row_counter_server-1)

# Calculate Number of VPS, Discount 1 because the items starts from A2
worksheet_Reports.write(number_of_vps, 1, row_counter_vps-1)

# Import Total cPanel Licenses Cost from cPanel_licenses sheet
worksheet_Reports.write_formula(total_cpanel_report, 1, '=cPanel_licenses!D8')

# Import Total Plesk Licenses Cost from Plesk_Licenses sheet
worksheet_Reports.write_formula(total_plesk_report, 1, '=Plesk_Licenses!D8')

# Import Total Windows Licenses Cost from Windows_Licenses sheet
worksheet_Reports.write_formula(total_windows_report, 1, '=Windows_Licenses!D8')

# Import Total Servers Cost from Servers_Pricing sheet
worksheet_Reports.write_formula(total_servers_report, 1, '=Servers_Pricing!L8')

# Import Total VPS Cost from VPS sheet
worksheet_Reports.write_formula(total_vps_report, 1, '=VPS!G8')

# Import Total Cabling from Cabling sheet
worksheet_Reports.write_formula(total_cabling_report, 1, '=Cabling!D8')

# Import Total Extra from Extras sheet
worksheet_Reports.write_formula(total_extras_report, 1, '=Extras!D8')

# Calculate the Total sum of Reports sheet
worksheet_Reports.write_formula(total_report_sum, 1,
            # GET the range to sum from the counter start and ending rows
            '=SUM(B%d:B%d)' % (total_cpanel_report+1, total_extras_report+1))

# Calculate the importing results
worksheet_Reports.write(1, 3,
            # GET the collected items from the global counters, Discount 1 from global counter which is for headers
                        # (['Service', 'ShPack', 'Description', 'From', 'To', 'Amount', 'EuroAmount')
            '%d items OF %d' % (global_counter_report, global_counter-1))

# Close the final XLSX file
workbook.close()

# Open the new export file, openpyxl has the magic to get every function work!
# If i didn't open that, XLSXWRITER will set the functions values as ZERO
wb1 = load_workbook(export_invoice)
wb1.save(export_invoice)

# Deleting tmp file
os.remove(export_temp)
